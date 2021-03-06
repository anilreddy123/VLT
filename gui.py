from Tkinter import *
import Tkinter
import subprocess
import ttk
import Tkinter as tk
from ScrolledText import *
import tkMessageBox
from openpyxl import load_workbook
from itertools import izip
import collections
import subprocess
import re

import sys
reload(sys)
sys.setdefaultencoding('utf-8')


class simpleapp_tk(Tk):
    def __init__(self,parent):
        Tk.__init__(self,parent)
        self.parent = parent
        
       
	workbook = load_workbook('Learning.xlsx', use_iterators=True)

	work_sheet=workbook.worksheets[0]
	
	print work_sheet
	res=[]

	keys=['id','name','eid','exp','skills','email']

	for i in range(2,work_sheet.max_row+1):
		d=collections.OrderedDict()
		#print d
		for k,j in izip(keys,range(1,work_sheet.max_column+1)):
		#for j in range(1,work_sheet.max_column+1):
			d[k]=work_sheet.cell(row=i,column=j).value
		
		#print work_sheet.cell(row=i,column=j).value

		res.append(d)
	for i in res:	
		if i['eid']==None:
			res.remove(i)	
	self.res=res
	self.initialize()

    def initialize(self):
        self.grid()

        self.frame = Frame(self, borderwidth=2, relief="sunken", width=1000, height=50)
        self.frame.grid(columnspan=3,row = 0)

        self.act = Label(self.frame,text = "Search with emp-id or Skill", fg = "red", font = 20)
        self.act.place(in_=self.frame, anchor="c", relx=.5, rely=.5)
        
        self.frame1 = Frame(self, borderwidth=2, relief="sunken", width = 300, height=550)
        self.frame1.grid(column = 0,row = 1, sticky = "WN")

        self.frame11 = Frame(self.frame1, borderwidth=1, relief="sunken", width = 300, height=40)
        self.frame11.grid(row=0)

        self.command = Label(self.frame11,text = "List", fg = "purple", font = 15)
        self.command.place(in_=self.frame11, anchor="c", relx=.5, rely=.5)

        self.frame12 = Frame(self.frame1, borderwidth=1, relief="sunken", width = 50, height=510)
        self.frame12.grid(row=1)
	self.tree = ttk.Treeview(self.frame12,height=25)

	for i,j in enumerate(self.res):
		self.tree.insert("",i,'VT-'+str(self.res[i]['eid']),text='VT-'+str(self.res[i]['eid']))
         
        vsb = ttk.Scrollbar(orient="vertical", command=self.tree.yview)
      
        self.tree.configure(yscrollcommand=vsb.set,)
	self.tree.grid(column=0, row=0, sticky='nsew', in_=self.frame12)
        vsb.grid(column=1, row=0, sticky='ns', in_=self.frame12)
        
	self.frame12.grid_columnconfigure(0, weight=3)
        self.frame12.grid_rowconfigure(0, weight=3)

	

	self.tree.bind("<<TreeviewSelect>>", self.On_tree_selected)
	
        self.frame2 = Frame(self, borderwidth=2, relief="sunken", width = 695, height=545)
        self.frame2.grid(column = 1,row = 1, sticky = "WN")

	self.frame21=Frame(self.frame2,borderwidth=1,relief="sunken",width=695,height=100)
	self.frame21.grid(row=0)
	#self.frame21.columnconfigure(0, weight=3)
	
	self.command_entry_id = Tkinter.Entry(self.frame21)
        self.command_entry_id.grid(row=0,column=1,padx=25, pady=5)
	self.commands2=Label(self.frame21,text="Employee ID",fg="purple",font=20)
	self.commands2.grid(row=0,column=0,padx=25, pady=5)
	
	self.commands = Button(self.frame21, text='send',fg="purple",font=10,command=self.on_button_id)
	self.commands.grid(row=0,column=2)

	self.command_entry_skill = Tkinter.Entry(self.frame21)
        self.command_entry_skill.grid(row=1,column=1,padx=25, pady=5)
	self.commands3=Label(self.frame21,text="Skill",fg="purple",font=20)
	self.commands3.grid(row=1,column=0,padx=25, pady=15)
	
	self.commands = Button(self.frame21, text='send',fg="purple",font=10,command=self.on_button)
	self.commands.grid(row=1,column=2)
	
	self.frame22=Frame(self.frame2,borderwidth=1,relief="sunken",width=695,height=445)
	self.frame22.grid(row=1)

	'''self.t=Text(self.frame22, borderwidth=1,bg='CadetBlue1',relief="sunken",width=695,height=445)
	self.t.place(in_=self.frame22)'''
	
	
	self.textPad=Frame(self.frame22)
        
	self.text=Text(self.textPad,bg='CadetBlue1', width=95,height=30)
         
        # add a vertical scroll bar to the text area
        self.scroll=Scrollbar(self.textPad,command=self.text.yview)
        self.text.configure(yscrollcommand=self.scroll.set)
         
        #pack everything
        self.text.pack(side=LEFT)
        self.scroll.pack(side=RIGHT,fill=Y)
        self.textPad.pack(side=TOP)
	
    	
        self.grid_columnconfigure(0,weight=1)
        self.grid_columnconfigure(1,weight=1)
        self.grid_columnconfigure(2,weight=1)

        self.grid_rowconfigure(0,weight=1)
        self.grid_rowconfigure(1,weight=1)
        self.grid_rowconfigure(2,weight=1)
       
        self.resizable(False,False)

   
	    
    def On_tree_selected(self, event):
	self.text.delete("1.0",END)
	if self.send > 0:
		self.text.delete("1.0",END)
		self.text.delete("1.0",END)
		#self.text3.delete("1.0",END)
	for item in self.tree.selection():
		self.send=0
		self.text.insert(END,item+'\n')
		for i in self.res:
			item=item.strip('VTvt-')
			match=re.search(re.escape(item),str(i['eid']).strip(' '),re.I)
			if match:
				line= "emp id  :VT-{0}\nname    :{1}\nemail   :{2}\nskills  :{3}\n\n".format(str(i['eid']),i['name'],i['email'],i['skills'])
				self.text.insert(END,line)
			
				break
	
		else:
			line= "No Data Found"
			self.text.insert(END,line)
		#print item
	#if (command[1]==item)
    def on_button_id(self):
	self.text.delete("1.0",END)
	if self.send > 0:
		self.text.delete("1.0",END)
		self.text.delete("1.0",END)
	
	self.emp_id=self.command_entry_id.get()
	
	
	if re.search(r'\d{3}',self.emp_id):
		self.emp_id=re.search(r'\d{3}',self.emp_id).group()
	
	else:
		self.text.insert(END,"Enter Valid Employee Id")
		return
	for i in self.res:
		#self.emp_id=self.emp_id.strip('VT-')
		match=re.search(re.escape(self.emp_id),str(i['eid']).strip(' '),re.I)
		if match:
			line= "emp id  :VT-{0}\nname    :{1}\nemail   :{2}\nskills  :{3}\n\n".format(str(i['eid']),i['name'],i['email'],i['skills'])
			self.text.insert(END,line)
			break
	
	else:
		line= "No Data Found"
		self.text.insert(END,line)
	
	
    def on_button(self):
	self.text.delete("1.0",END)
	if self.send > 0:
		self.text.delete("1.0",END)
		self.text.delete("1.0",END)
	self.sk = ''
	self.sk=self.command_entry_skill.get()
	
	if ',' in self.sk:
		self.sk=self.sk.split(',')
	elif ' ' in self.sk:
		self.sk=self.sk.split(' ')
	elif self.sk == '':
		return self.text.insert(END,"Enter Any Skill")
		
	else:
		self.sk = [self.sk]
	list_id=collections.OrderedDict()
	print self.sk
	names=[]
	for j in self.sk:
		for i in self.res:
			
			match=re.search(r'\b'+re.escape(j)+r'\b',i['skills'],re.I)

			if match:
				
				ide='VT-'+str(i['eid']).replace(' ','')

				
				if list_id.has_key(ide):
					
					list_id[ide]['skills'].append(j)	
					
				else:

					ide='VT-'+str(i['eid']).replace(' ','')
					list_id[ide]={}
					list_id[ide]['skills']=[j]
					list_id[ide]['name']=i['name']
					list_id[ide]['email']=i['email']
					list_id[ide]['exp']=str(i['exp'])
				
	if len(list_id)>0:
		val=1
		for i,k in list_id.iteritems():
			text= "emp id  :{0:^6}\nName    :{1:<20}\nEmail   :{2:<40}\nskills  :{3:<50}\n\n".format(str(i),k['name'],k['email'],','.join(k['skills']))
			#print text
			self.text.insert(END,str(val)+'\n')
			self.text.insert(END,text)
			val+=1
	else:
		self.text.insert(END,'No Data Found For Skills')


if __name__ == "__main__":
    app = simpleapp_tk(None)
    app.title('ACT')
    app.geometry("1000x600")
    app.mainloop()
